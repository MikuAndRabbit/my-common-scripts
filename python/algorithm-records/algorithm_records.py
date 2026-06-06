"""Parser for ``Algorithm Records.xlsx``.

The workbook is a刷题记录 table organised as one sheet per平台 (Acwing /
LeetCode / LC-Special / 面试题 / 牛客) plus a颜色备注 sheet that documents
the cell-background → 难度评价 mapping.

每个数据 sheet 的列结构如下::

    | 题目序号 | 题目名称 |  1  |  2  |  3  | ...  N  |
    | -------- | -------- | --- | --- | --- | ------- |
    | 1        | 两数之和 | 日期 | 日期 |     |         |

* 第 N 列的表头是整数 ``N``, 代表第 N 次刷题;
* 单元格的值是该次刷题的日期 (``datetime``);
* 单元格的背景颜色编码了刷题的体感难度 (见 ``颜色备注`` sheet):

    - ``FFADD88D`` (浅绿) → 流畅
    - ``FFF4B382`` (橙)   → 卡顿
    - ``FFEF949F`` (粉红) → 困难

  颜色既可能是直接的 RGB, 也可能是 ``theme + tint`` 形式 (WPS 用调色板填色
  时常见). 解析器会把 theme color 通过 workbook 自带的 ``xl/theme/theme1.xml``
  解算成实际 RGB, 再按欧氏距离与图例做最近邻匹配, 因此两种填色方式都能识别.

本模块把整个 workbook 解析成一组 ``Problem`` / ``Attempt`` dataclass, 并提供
查询/统计/导出 (JSON, dict) 的便捷方法. 既可作为库使用, 也可直接 ``python
algorithm_records.py`` 当 CLI 跑.
"""

from __future__ import annotations

import argparse
import colorsys
import json
import sys
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# 常量 / 类型
# ---------------------------------------------------------------------------

#: 颜色备注 sheet 的名字 — 不参与题目解析.
LEGEND_SHEET = "颜色备注"

#: 颜色 (大写, 含 alpha) → 难度评价. 来自 ``颜色备注`` sheet,
#: 在 :func:`load_color_legend` 中会用文件里读到的值覆盖.
DEFAULT_DIFFICULTY_BY_COLOR: dict[str, str] = {
    "FFADD88D": "流畅",
    "FFF4B382": "卡顿",
    "FFEF949F": "困难",
}

#: 没有背景色 / 未在图例中的颜色, 统一记作 ``未标注``.
UNLABELED = "未标注"

#: OOXML theme color 索引 → ``<a:clrScheme>`` 子元素名. 注意 0/1 和 2/3 是
#: 反着的 — 这是 OOXML 的著名坑.
_THEME_INDEX_TO_ELEMENT = {
    0: "lt1",
    1: "dk1",
    2: "lt2",
    3: "dk2",
    4: "accent1",
    5: "accent2",
    6: "accent3",
    7: "accent4",
    8: "accent5",
    9: "accent6",
    10: "hlink",
    11: "folHlink",
}

#: WPS / Office 默认主题色 (theme1.xml 里 ``<a:clrScheme>`` 各 accent 的 RGB).
#: workbook 自带的 theme1.xml 找不到时回落用. 当前文件确实用了 WPS 主题:
#: accent2=橙 (EE822F), accent4=绿 (75BD42), accent6=红 (E54C5E) —
#: 加上 tint=0.4 之后正好等于 ``颜色备注`` 里那 3 个 RGB.
_FALLBACK_THEME_RGB: dict[int, str] = {
    0: "FFFFFF", 1: "000000", 2: "E7E6E6", 3: "44546A",
    4: "4874CB", 5: "EE822F", 6: "F2BA02", 7: "75BD42",
    8: "30C0B4", 9: "E54C5E", 10: "0026E5", 11: "7E1FAD",
}

#: 把 (theme, tint) → 6 位 RGB hex 算出来后, 与图例 RGB 做最近邻匹配时的
#: 最大欧氏距离 (R/G/B 各 0..255). 实测 WPS 把 ``EE822F`` + tint 0.4 算成
#: ``F5B482``, 而图例是 ``F4B382``, 距离 ≈ 1.7 — 阈值 30 足够稳健, 同时
#: 也不会把绿/橙/红互相误吃 (最近的两色距离 > 100).
_COLOR_MATCH_THRESHOLD = 30.0


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Attempt:
    """一次刷题记录."""

    #: 第几次刷此题 (1-based, 与表头列号一致).
    index: int
    #: 刷题日期.
    date: date
    #: 体感难度 (流畅/卡顿/困难/未标注).
    difficulty: str
    #: 原始背景色 (大写 8 位 ARGB), 没有填充时为 ``None``.
    color: str | None

    def to_dict(self) -> dict:
        d = asdict(self)
        d["date"] = self.date.isoformat()
        return d


@dataclass
class Problem:
    """一道题及其所有刷题记录."""

    #: 来源 sheet 名 (Acwing / LeetCode / ...).
    source: str
    #: 题目序号 (Acwing 为 ``int``; 其它平台多为带前缀的 ``str``, e.g. ``LCP 01``).
    problem_id: int | str
    #: 题目名称, 占位行会是 ``None``.
    name: str | None
    #: 所有刷题记录, 按 ``index`` 升序.
    attempts: list[Attempt] = field(default_factory=list)

    @property
    def attempt_count(self) -> int:
        return len(self.attempts)

    @property
    def first_attempt(self) -> Attempt | None:
        return self.attempts[0] if self.attempts else None

    @property
    def last_attempt(self) -> Attempt | None:
        return self.attempts[-1] if self.attempts else None

    @property
    def is_attempted(self) -> bool:
        return bool(self.attempts)

    def to_dict(self) -> dict:
        return {
            "source": self.source,
            "problem_id": self.problem_id,
            "name": self.name,
            "attempt_count": self.attempt_count,
            "attempts": [a.to_dict() for a in self.attempts],
        }


@dataclass
class AlgorithmRecords:
    """整本 workbook 的解析结果."""

    #: 文件路径.
    path: Path
    #: 颜色 → 难度的图例 (从 workbook 读出, 找不到则回落到默认).
    color_legend: dict[str, str]
    #: 平台名 → 该平台的所有题目列表 (保持原 sheet 顺序).
    problems_by_source: dict[str, list[Problem]]

    # ---- 便捷访问 ------------------------------------------------------

    @property
    def sources(self) -> list[str]:
        return list(self.problems_by_source)

    def all_problems(self) -> Iterator[Problem]:
        """按平台顺序遍历全部题目 (含未做)."""
        for problems in self.problems_by_source.values():
            yield from problems

    def attempted_problems(self) -> Iterator[Problem]:
        """只遍历至少做过一次的题目."""
        return (p for p in self.all_problems() if p.is_attempted)

    def all_attempts(self) -> Iterator[tuple[Problem, Attempt]]:
        """遍历所有刷题事件, 附带其所属题目."""
        for problem in self.attempted_problems():
            for attempt in problem.attempts:
                yield problem, attempt

    # ---- 查询 ----------------------------------------------------------

    def find(self, source: str, problem_id: int | str) -> Problem | None:
        """按平台 + 题号查找; 找不到返回 ``None``."""
        for p in self.problems_by_source.get(source, ()):
            if p.problem_id == problem_id:
                return p
        return None

    def search_by_name(
        self, keyword: str, *, source: str | None = None
    ) -> list[Problem]:
        """按题名子串模糊查找 (大小写不敏感)."""
        kw = keyword.lower()
        pool: Iterable[Problem]
        pool = (
            self.problems_by_source.get(source, ())
            if source is not None
            else self.all_problems()
        )
        return [p for p in pool if p.name and kw in p.name.lower()]

    # ---- 统计 ----------------------------------------------------------

    def summary(self) -> dict[str, dict[str, int]]:
        """按平台统计: 总题数 / 已做 / 总刷题次数 / 各难度计数."""
        result: dict[str, dict[str, int]] = {}
        for source, problems in self.problems_by_source.items():
            attempted = [p for p in problems if p.is_attempted]
            difficulty_counter: Counter[str] = Counter()
            attempt_count = 0
            for p in attempted:
                attempt_count += p.attempt_count
                for a in p.attempts:
                    difficulty_counter[a.difficulty] += 1
            result[source] = {
                "total_problems": len(problems),
                "attempted_problems": len(attempted),
                "total_attempts": attempt_count,
                **{f"难度_{k}": v for k, v in difficulty_counter.items()},
            }
        return result

    def daily_counts(
        self, *, source: str | None = None
    ) -> dict[date, int]:
        """按日期统计刷题次数 (跨平台或单平台)."""
        counter: Counter[date] = Counter()
        problems = (
            self.problems_by_source.get(source, ())
            if source is not None
            else self.all_problems()
        )
        for p in problems:
            for a in p.attempts:
                counter[a.date] += 1
        return dict(sorted(counter.items()))

    # ---- 导出 ----------------------------------------------------------

    def to_dict(self) -> dict:
        return {
            "path": str(self.path),
            "color_legend": self.color_legend,
            "summary": self.summary(),
            "problems_by_source": {
                src: [p.to_dict() for p in probs]
                for src, probs in self.problems_by_source.items()
            },
        }

    def to_json(self, *, indent: int | None = 2) -> str:
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=indent)


# ---------------------------------------------------------------------------
# 解析逻辑
# ---------------------------------------------------------------------------


def _read_theme_rgbs(wb: Workbook) -> dict[int, str]:
    """从 workbook 的 ``xl/theme/theme1.xml`` 里读出每个 theme 索引的 RGB.

    解析失败时返回 :data:`_FALLBACK_THEME_RGB` (WPS 默认主题).
    openpyxl 把 theme xml 字节流存在 ``wb.loaded_theme`` 里 (是 bytes 或 str).
    """
    raw = getattr(wb, "loaded_theme", None)
    if raw is None:
        return dict(_FALLBACK_THEME_RGB)
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8", errors="replace")
    try:
        import xml.etree.ElementTree as ET
        root = ET.fromstring(raw)
        # 命名空间是固定的 a=http://schemas.openxmlformats.org/drawingml/2006/main
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        scheme = root.find(".//a:clrScheme", ns)
        if scheme is None:
            return dict(_FALLBACK_THEME_RGB)
        out: dict[int, str] = {}
        for idx, name in _THEME_INDEX_TO_ELEMENT.items():
            el = scheme.find(f"a:{name}", ns)
            if el is None:
                continue
            srgb = el.find("a:srgbClr", ns)
            if srgb is not None and "val" in srgb.attrib:
                out[idx] = srgb.attrib["val"].upper()
                continue
            sysclr = el.find("a:sysClr", ns)
            if sysclr is not None and "lastClr" in sysclr.attrib:
                out[idx] = sysclr.attrib["lastClr"].upper()
        # 缺的索引用 fallback 补全, 防止意外 KeyError
        for idx, rgb in _FALLBACK_THEME_RGB.items():
            out.setdefault(idx, rgb)
        return out
    except Exception:
        return dict(_FALLBACK_THEME_RGB)


def _apply_tint(rgb_hex: str, tint: float) -> str:
    """对 6 位 RGB hex 应用 OOXML tint, 返回 6 位 hex.

    OOXML 规则: 把 RGB 转 HLS, 调整 L 通道:
        tint < 0:  L' = L * (1 + tint)
        tint > 0:  L' = L * (1 - tint) + (1 - (1 - tint))
                       = L * (1 - tint) + tint
    再转回 RGB. tint=0 (或 None) 时原样返回.
    """
    if not tint:
        return rgb_hex.upper()
    r = int(rgb_hex[0:2], 16) / 255
    g = int(rgb_hex[2:4], 16) / 255
    b = int(rgb_hex[4:6], 16) / 255
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1 + tint)
    else:
        l = l * (1 - tint) + tint
    l = max(0.0, min(1.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return f"{int(round(r*255)):02X}{int(round(g*255)):02X}{int(round(b*255)):02X}"


class _ColorClassifier:
    """把任意 cell 的填充 → ``(规范化色 key, 难度 label)``.

    - 处理 3 种填充颜色来源: RGB / theme+tint / indexed (后者罕见, 忽略).
    - 用 ``颜色备注`` sheet 给出的 RGB 图例做最近邻匹配 (≤
      :data:`_COLOR_MATCH_THRESHOLD`).
    - 把已经匹配过的 cell 颜色缓存起来, 避免 reparse.
    """

    def __init__(
        self,
        legend_rgb: dict[str, str],
        theme_rgbs: dict[int, str],
        threshold: float = _COLOR_MATCH_THRESHOLD,
    ) -> None:
        self.theme_rgbs = theme_rgbs
        self.threshold = threshold
        # legend_rgb key 可能带 alpha (8 位) 也可能不带 (6 位), 统一成 6 位.
        self._legend: list[tuple[str, str, tuple[int, int, int]]] = []
        for color_key, label in legend_rgb.items():
            rgb6 = color_key[-6:].upper()
            self._legend.append((color_key, label, _hex_to_rgb(rgb6)))
        self._cache: dict[str, tuple[str | None, str]] = {}

    def classify(self, cell: Cell) -> tuple[str | None, str]:
        """返回 ``(色 key, 难度)``. 没填色或匹不到时返回 ``(None, '未标注')``
        或 ``(原始 key, '未标注')``.
        """
        key = _cell_color_key(cell, self.theme_rgbs)
        if key is None:
            return (None, UNLABELED)
        cached = self._cache.get(key)
        if cached is not None:
            return cached

        # key 形如 'rgb:FFADD88D' 或 'resolved:F5B482' (theme 解算后) 或
        # 'theme:7/0.4' (theme 解算失败时的兜底). 取出能比较的 RGB 6 位.
        rgb6: str | None = None
        if key.startswith("rgb:"):
            rgb6 = key[4:][-6:].upper()
        elif key.startswith("resolved:"):
            rgb6 = key[len("resolved:"):].upper()

        label = UNLABELED
        if rgb6 is not None:
            target = _hex_to_rgb(rgb6)
            best_dist = float("inf")
            best_label = UNLABELED
            for _, cand_label, cand_rgb in self._legend:
                d = _rgb_distance(target, cand_rgb)
                if d < best_dist:
                    best_dist = d
                    best_label = cand_label
            if best_dist <= self.threshold:
                label = best_label

        result = (key, label)
        self._cache[key] = result
        return result


def _hex_to_rgb(hex6: str) -> tuple[int, int, int]:
    return (int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16))


def _rgb_distance(a: tuple[int, int, int], b: tuple[int, int, int]) -> float:
    return ((a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2 + (a[2] - b[2]) ** 2) ** 0.5


def _cell_color_key(cell: Cell, theme_rgbs: dict[int, str]) -> str | None:
    """把 cell 的填充规范化成一个稳定的字符串 key (或 ``None``).

    - 实心 RGB:    ``"rgb:FFADD88D"`` (大写, 保留 alpha 如果有)
    - 实心 theme:  优先算出最终 RGB, 返回 ``"resolved:RRGGBB"``;
                  算不出时返回 ``"theme:<idx>/<tint>"`` 作为兜底
    - 非实心 / 没填:  ``None``
    """
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "rgb" and isinstance(fg.rgb, str):
        return f"rgb:{fg.rgb.upper()}"
    if fg.type == "theme":
        base = theme_rgbs.get(int(fg.theme))
        if base:
            return f"resolved:{_apply_tint(base, float(fg.tint or 0))}"
        return f"theme:{fg.theme}/{fg.tint}"
    return None


def _cell_color(cell: Cell, theme_rgbs: dict[int, str] | None = None) -> str | None:
    """向后兼容的 wrapper — 返回 ``Attempt.color`` 字段使用的字符串.

    保留旧行为: RGB 单元格直接返回 6/8 位 hex (大写); theme 单元格返回
    ``"theme:<idx>/<tint>"``. 想要解算后的 RGB 请用 :class:`_ColorClassifier`.
    """
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "rgb" and isinstance(fg.rgb, str):
        return fg.rgb.upper()
    if fg.type == "theme":
        if theme_rgbs is not None:
            base = theme_rgbs.get(int(fg.theme))
            if base:
                return _apply_tint(base, float(fg.tint or 0))
        return f"theme:{fg.theme}/{fg.tint}"
    return None


def _coerce_date(value: object) -> date | None:
    """把单元格里的日期值正规化成 ``date``.

    支持: ``datetime`` (openpyxl 默认), ``date``, ``YYYY-MM-DD`` 字符串.
    其它类型返回 ``None`` (调用方会跳过).
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def load_color_legend(path: str | Path) -> dict[str, str]:
    """从 workbook 的 ``颜色备注`` sheet 解析 颜色→难度 映射.

    缺失或异常时回落到 :data:`DEFAULT_DIFFICULTY_BY_COLOR`, 不会抛.
    """
    wb = load_workbook(path, data_only=True)
    theme_rgbs = _read_theme_rgbs(wb)
    if LEGEND_SHEET not in wb.sheetnames:
        return dict(DEFAULT_DIFFICULTY_BY_COLOR)
    ws = wb[LEGEND_SHEET]
    legend: dict[str, str] = {}
    # 跳过表头行
    for row in ws.iter_rows(min_row=2):
        if len(row) < 2:
            continue
        color_cell, label_cell = row[0], row[1]
        label = label_cell.value
        color = _cell_color(color_cell, theme_rgbs)
        if color and isinstance(label, str) and label.strip():
            legend[color] = label.strip()
    return legend or dict(DEFAULT_DIFFICULTY_BY_COLOR)


def _parse_sheet(
    ws: Worksheet, *, classifier: "_ColorClassifier"
) -> list[Problem]:
    """解析单个平台 sheet, 返回该 sheet 的全部 ``Problem``."""
    rows = ws.iter_rows()
    try:
        header = next(rows)
    except StopIteration:
        return []

    # 表头形如 ('题目序号', '题目名称', 1, 2, 3, ...). 我们需要知道每一列
    # 对应第几次刷题 — 直接取表头里的整数. 若某列表头为空 (Acwing 末尾有 2
    # 个空列), 该列直接跳过.
    attempt_columns: list[tuple[int, int]] = []  # (col_index_in_row, attempt_index)
    for col_idx, cell in enumerate(header):
        if col_idx < 2:
            continue  # 题目序号 / 题目名称
        val = cell.value
        if isinstance(val, int):
            attempt_columns.append((col_idx, val))
        elif isinstance(val, str) and val.strip().isdigit():
            attempt_columns.append((col_idx, int(val.strip())))
        # else: 跳过 (空表头列)

    problems: list[Problem] = []
    for row in rows:
        # 空行 (id 和 name 都没有) 跳过.
        if row[0].value is None and row[1].value is None:
            continue
        problem_id = row[0].value
        if problem_id is None:
            continue
        name = row[1].value if len(row) > 1 else None
        if isinstance(name, str):
            name = name.strip() or None

        attempts: list[Attempt] = []
        for col_idx, attempt_idx in attempt_columns:
            if col_idx >= len(row):
                break
            cell = row[col_idx]
            d = _coerce_date(cell.value)
            if d is None:
                continue
            color_key, difficulty = classifier.classify(cell)
            # Attempt.color 保留对外友好的字符串: rgb: 前缀去掉只留 hex,
            # resolved: 也去前缀, theme: 保留原样.
            color: str | None
            if color_key is None:
                color = None
            elif color_key.startswith("rgb:"):
                color = color_key[4:]
            elif color_key.startswith("resolved:"):
                color = color_key[len("resolved:"):]
            else:
                color = color_key
            attempts.append(
                Attempt(
                    index=attempt_idx,
                    date=d,
                    difficulty=difficulty,
                    color=color,
                )
            )
        attempts.sort(key=lambda a: a.index)
        problems.append(
            Problem(
                source=ws.title,
                problem_id=problem_id,
                name=name,
                attempts=attempts,
            )
        )
    return problems


def parse(path: str | Path) -> AlgorithmRecords:
    """解析整本 ``Algorithm Records.xlsx``.

    Parameters
    ----------
    path:
        xlsx 文件路径.

    Returns
    -------
    :class:`AlgorithmRecords`
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件: {path}")
    wb = load_workbook(path, data_only=True)
    theme_rgbs = _read_theme_rgbs(wb)
    # 图例: 重新走一遍 (load_color_legend 内部又开了 workbook, 这里我们已经
    # 有 wb, 直接复用更省事).
    color_legend: dict[str, str] = {}
    if LEGEND_SHEET in wb.sheetnames:
        ws = wb[LEGEND_SHEET]
        for row in ws.iter_rows(min_row=2):
            if len(row) < 2:
                continue
            color = _cell_color(row[0], theme_rgbs)
            label = row[1].value
            if color and isinstance(label, str) and label.strip():
                color_legend[color] = label.strip()
    if not color_legend:
        color_legend = dict(DEFAULT_DIFFICULTY_BY_COLOR)
    classifier = _ColorClassifier(color_legend, theme_rgbs)

    problems_by_source: dict[str, list[Problem]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == LEGEND_SHEET:
            continue
        ws = wb[sheet_name]
        problems_by_source[sheet_name] = _parse_sheet(ws, classifier=classifier)
    return AlgorithmRecords(
        path=path,
        color_legend=color_legend,
        problems_by_source=problems_by_source,
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _cmd_summary(records: AlgorithmRecords) -> None:
    summary = records.summary()
    width = max(len(s) for s in summary) if summary else 0
    print(f"文件: {records.path}")
    print(f"颜色图例: {records.color_legend}")
    print()
    header = f"{'平台'.ljust(width)}  总题数  已做题  总次数  流畅  卡顿  困难  未标注"
    print(header)
    print("-" * len(header))
    for source, s in summary.items():
        print(
            f"{source.ljust(width)}  "
            f"{s['total_problems']:>6}  "
            f"{s['attempted_problems']:>6}  "
            f"{s['total_attempts']:>6}  "
            f"{s.get('难度_流畅', 0):>4}  "
            f"{s.get('难度_卡顿', 0):>4}  "
            f"{s.get('难度_困难', 0):>4}  "
            f"{s.get('难度_未标注', 0):>6}"
        )


def _cmd_find(records: AlgorithmRecords, source: str, pid: str) -> None:
    # 题号可能是 int (Acwing/LeetCode) 也可能是 str (其它平台). 尝试两种.
    candidates: list[int | str] = [pid]
    try:
        candidates.append(int(pid))
    except ValueError:
        pass
    for cand in candidates:
        p = records.find(source, cand)
        if p:
            print(json.dumps(p.to_dict(), ensure_ascii=False, indent=2))
            return
    print(f"未找到: source={source} id={pid}", file=sys.stderr)
    sys.exit(1)


def _cmd_search(records: AlgorithmRecords, keyword: str, source: str | None) -> None:
    hits = records.search_by_name(keyword, source=source)
    if not hits:
        print("无匹配", file=sys.stderr)
        sys.exit(1)
    for p in hits:
        last = p.last_attempt
        last_str = (
            f"最近 {last.date.isoformat()} ({last.difficulty})"
            if last
            else "未做"
        )
        print(f"[{p.source}] {p.problem_id}\t{p.name}\t次数={p.attempt_count}\t{last_str}")


def _cmd_export(records: AlgorithmRecords, out_path: str | None) -> None:
    payload = records.to_json()
    if out_path:
        Path(out_path).write_text(payload, encoding="utf-8")
        print(f"已写入 {out_path} ({len(payload)} bytes)")
    else:
        print(payload)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="解析 Algorithm Records.xlsx 的命令行工具",
    )
    parser.add_argument(
        "--file",
        "-f",
        default="Algorithm Records.xlsx",
        help="xlsx 文件路径 (默认: ./Algorithm Records.xlsx)",
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    sub.add_parser("summary", help="打印各平台的统计概览")

    p_find = sub.add_parser("find", help="按平台+题号精确查找")
    p_find.add_argument("source", help="平台名, e.g. LeetCode / Acwing / 牛客")
    p_find.add_argument("id", help="题目序号, e.g. 1 或 'LCP 01'")

    p_search = sub.add_parser("search", help="按题名子串模糊搜索")
    p_search.add_argument("keyword", help="关键字 (大小写不敏感)")
    p_search.add_argument(
        "--source", "-s", default=None, help="可选: 限定平台"
    )

    p_export = sub.add_parser("export", help="把全部解析结果导出成 JSON")
    p_export.add_argument(
        "--out", "-o", default=None, help="输出文件路径; 缺省时打到 stdout"
    )

    args = parser.parse_args(argv)
    records = parse(args.file)

    if args.cmd == "summary":
        _cmd_summary(records)
    elif args.cmd == "find":
        _cmd_find(records, args.source, args.id)
    elif args.cmd == "search":
        _cmd_search(records, args.keyword, args.source)
    elif args.cmd == "export":
        _cmd_export(records, args.out)


if __name__ == "__main__":
    main()
